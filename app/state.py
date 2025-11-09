import reflex as rx
from typing import TypedDict
import openpyxl
import datetime
import logging


class Item(TypedDict):
    id: int
    name: str
    sku: str
    quantity: int
    category: str
    unit_price_naira: float | None
    unit_price_dollar: float | None
    unit_price_euro: float | None
    image_url: str


class InflowTransaction(TypedDict):
    id: int
    item_name: str
    quantity: int
    supplier: str
    date: str
    notes: str


class OutflowTransaction(TypedDict):
    id: int
    item_name: str
    quantity: int
    destination: str
    date: str
    notes: str


class NavItem(TypedDict):
    name: str
    icon: str
    href: str


class InventoryState(rx.State):
    """Manages the state for the inventory management system."""

    current_page: str = "Portfolio"
    portfolio_search: str = ""
    inflow_search: str = ""
    outflow_search: str = ""
    items: list[Item] = []
    nav_items: list[NavItem] = [
        {"name": "Portfolio", "icon": "package", "href": "/"},
        {"name": "Inflow", "icon": "arrow-down-circle", "href": "/inflow"},
        {"name": "Outflow", "icon": "arrow-up-circle", "href": "/outflow"},
        {"name": "Admin", "icon": "sliders-horizontal", "href": "/admin"},
    ]
    inflow_transactions: list[InflowTransaction] = []
    outflow_transactions: list[OutflowTransaction] = []

    @rx.event
    def on_load(self):
        """Load data from the Excel file when the app starts."""
        if not self.items:
            self._load_excel_data()

    def _load_excel_data(self):
        try:
            wb = openpyxl.load_workbook(
                "assets/INVENTORY STOCK SHEEETS 2025 (1).xlsx", data_only=True
            )
            sheet_items = wb["2025 stock sheet(NEW)"]
            loaded_items = []
            for i, row in enumerate(sheet_items.iter_rows(min_row=2, values_only=True)):
                if row[0] is not None and row[1] is not None:
                    try:
                        sku = str(row[0])
                        quantity = (
                            int(row[5]) if isinstance(row[5], (int, float)) else 0
                        )
                        loaded_items.append(
                            {
                                "id": i + 1,
                                "name": str(row[1]),
                                "sku": sku,
                                "quantity": quantity,
                                "category": "General",
                                "unit_price_naira": float(row[2])
                                if isinstance(row[2], (int, float))
                                else None,
                                "unit_price_dollar": float(row[3])
                                if isinstance(row[3], (int, float))
                                else None,
                                "unit_price_euro": float(row[4])
                                if isinstance(row[4], (int, float))
                                else None,
                                "image_url": "/placeholder.svg",
                            }
                        )
                    except (ValueError, TypeError) as e:
                        logging.exception(
                            f"Skipping item row due to parsing error: {e} | Row data: {row}"
                        )
            self.items = loaded_items
            sheet_inflow = wb["INFLOW"]
            loaded_inflow = []
            for row in sheet_inflow.iter_rows(min_row=4, values_only=True):
                if (
                    row[1] is not None
                    and row[3] is not None
                    and isinstance(row[1], int)
                ):
                    try:
                        date_val = (
                            row[0].strftime("%Y-%m-%d")
                            if isinstance(row[0], datetime.datetime)
                            else str(row[0])
                        )
                        notes = (
                            f"Unit Price: {row[6]}, Total: {row[7]}"
                            if row[6] is not None and row[7] is not None
                            else ""
                        )
                        loaded_inflow.append(
                            {
                                "id": int(row[1]),
                                "item_name": str(row[3]),
                                "quantity": int(row[5]) if row[5] is not None else 0,
                                "supplier": str(row[4]) if row[4] else "N/A",
                                "date": date_val,
                                "notes": notes,
                            }
                        )
                    except (ValueError, TypeError) as e:
                        logging.exception(
                            f"Skipping inflow row due to parsing error: {e} | Row data: {row}"
                        )
            self.inflow_transactions = loaded_inflow
            sheet_outflow = wb["OUTFLOW"]
            loaded_outflow = []
            for row in sheet_outflow.iter_rows(min_row=4, values_only=True):
                if (
                    row[1] is not None
                    and row[3] is not None
                    and isinstance(row[1], int)
                ):
                    try:
                        date_val = (
                            row[0].strftime("%Y-%m-%d")
                            if isinstance(row[0], datetime.datetime)
                            else str(row[0])
                        )
                        notes = (
                            f"Stock remaining: {row[6]}" if row[6] is not None else ""
                        )
                        loaded_outflow.append(
                            {
                                "id": int(row[1]),
                                "item_name": str(row[3]),
                                "quantity": int(row[5]) if row[5] is not None else 0,
                                "destination": str(row[4]) if row[4] else "N/A",
                                "date": date_val,
                                "notes": notes,
                            }
                        )
                    except (ValueError, TypeError) as e:
                        logging.exception(
                            f"Skipping outflow row due to parsing error: {e} | Row data: {row}"
                        )
            self.outflow_transactions = loaded_outflow
        except FileNotFoundError as e:
            logging.exception(f"Error: Inventory Excel file not found. {e}")
        except Exception as e:
            logging.exception(
                f"An unexpected error occurred while loading Excel data: {e}"
            )

    @rx.event
    def set_page(self, page_name: str):
        self.current_page = page_name

    @rx.event
    def add_item(self, form_data: dict):
        """Adds a new item to the inventory from form data."""
        new_id = max((item["id"] for item in self.items)) + 1 if self.items else 1
        new_item: Item = {
            "id": new_id,
            "name": form_data["name"],
            "sku": form_data["sku"],
            "quantity": int(form_data["quantity"]),
            "category": form_data["category"],
            "unit_price_naira": float(form_data["unit_price_naira"])
            if form_data.get("unit_price_naira")
            else None,
            "unit_price_dollar": float(form_data["unit_price_dollar"])
            if form_data.get("unit_price_dollar")
            else None,
            "unit_price_euro": float(form_data["unit_price_euro"])
            if form_data.get("unit_price_euro")
            else None,
            "image_url": "/placeholder.svg",
        }
        self.items.append(new_item)

    @rx.event
    def delete_item(self, item_id: int):
        """Deletes an item from the inventory."""
        self.items = [item for item in self.items if item["id"] != item_id]

    @rx.event
    def add_inflow_transaction(self, form_data: dict):
        """Adds a new inflow transaction and updates item quantity."""
        new_id = (
            max((t["id"] for t in self.inflow_transactions)) + 1
            if self.inflow_transactions
            else 1
        )
        new_transaction: InflowTransaction = {
            "id": new_id,
            "item_name": form_data["item_name"],
            "quantity": int(form_data["quantity"]),
            "supplier": form_data["supplier"],
            "date": form_data["date"],
            "notes": form_data["notes"],
        }
        self.inflow_transactions.append(new_transaction)
        for i, item in enumerate(self.items):
            if item["name"] == new_transaction["item_name"]:
                self.items[i]["quantity"] += new_transaction["quantity"]
                break

    @rx.event
    def add_outflow_transaction(self, form_data: dict):
        """Adds a new outflow transaction and updates item quantity."""
        new_id = (
            max((t["id"] for t in self.outflow_transactions)) + 1
            if self.outflow_transactions
            else 1
        )
        new_transaction: OutflowTransaction = {
            "id": new_id,
            "item_name": form_data["item_name"],
            "quantity": int(form_data["quantity"]),
            "destination": form_data["destination"],
            "date": form_data["date"],
            "notes": form_data["notes"],
        }
        self.outflow_transactions.append(new_transaction)
        for i, item in enumerate(self.items):
            if item["name"] == new_transaction["item_name"]:
                self.items[i]["quantity"] -= new_transaction["quantity"]
                break

    @rx.var
    def filtered_items(self) -> list[Item]:
        """Returns filtered items based on the search query."""
        search_lower = self.portfolio_search.lower()
        if not search_lower:
            return self.items
        return [
            item
            for item in self.items
            if search_lower in item["name"].lower()
            or search_lower in item["category"].lower()
        ]

    @rx.var
    def filtered_inflow(self) -> list[InflowTransaction]:
        """Returns filtered inflow transactions based on the search query."""
        search_lower = self.inflow_search.lower()
        if not search_lower:
            return self.inflow_transactions
        return [
            t
            for t in self.inflow_transactions
            if search_lower in t["item_name"].lower()
            or search_lower in t["supplier"].lower()
        ]

    @rx.var
    def filtered_outflow(self) -> list[OutflowTransaction]:
        """Returns filtered outflow transactions based on the search query."""
        search_lower = self.outflow_search.lower()
        if not search_lower:
            return self.outflow_transactions
        return [
            t
            for t in self.outflow_transactions
            if search_lower in t["item_name"].lower()
            or search_lower in t["destination"].lower()
        ]

    @rx.var
    def total_inflow_quantity(self) -> int:
        """Returns the total quantity from all inflow transactions."""
        return sum((t["quantity"] for t in self.inflow_transactions))

    @rx.var
    def total_outflow_quantity(self) -> int:
        """Returns the total quantity from all outflow transactions."""
        return sum((t["quantity"] for t in self.outflow_transactions))